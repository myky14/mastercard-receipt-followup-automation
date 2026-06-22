"""Excel and ZIP export helpers."""

from __future__ import annotations

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import CARDHOLDER_NAMES, OUTPUT_COLUMNS


CARDHOLDER_OUTPUT_COLUMNS = [
    "Date",
    "Bank description",
    "From/To",
    "Amount",
    "Card number",
    "Cardholder Name",
]


def build_output_frames(results_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Split matched results into the required workbook outputs."""
    frames: dict[str, pd.DataFrame] = {}
    cardholder_ready = results_df["Match confidence"].isin(["High", "Medium", "Manual"])

    for cardholder_name in CARDHOLDER_NAMES:
        filename = cardholder_output_filename(cardholder_name)
        frames[filename] = build_cardholder_output_frame(
            results_df.loc[cardholder_ready & (results_df["Cardholder name"] == cardholder_name)]
        )

    frames["Matching_Log.xlsx"] = _ensure_output_columns(results_df)
    frames["Need_Review.xlsx"] = _ensure_output_columns(
        results_df.loc[results_df["Match confidence"] == "Review"]
    )
    frames["Unmatched_QBO.xlsx"] = _ensure_output_columns(
        results_df.loc[results_df["Match confidence"] == "Unmatched"]
    )

    return frames


def cardholder_output_filename(cardholder_name: str) -> str:
    """Build the manager-ready missing receipt workbook filename."""
    return f"{cardholder_name.replace(' ', '_')}_missing_receipts.xlsx"


def build_cardholder_output_frame(results_df: pd.DataFrame) -> pd.DataFrame:
    """Build clean manager-ready cardholder output columns."""
    output = pd.DataFrame(
        {
            "Date": results_df["QBO Date"],
            "Bank description": results_df["QBO Bank description"],
            "From/To": results_df["QBO From/To"],
            "Amount": results_df.apply(_qbo_signed_amount, axis=1),
            "Card number": results_df["Card number"],
            "Cardholder Name": results_df["Cardholder name"],
        }
    )
    return output.reindex(columns=CARDHOLDER_OUTPUT_COLUMNS)


def create_output_zip(results_df: pd.DataFrame) -> BytesIO:
    """Create the required workbook ZIP in memory for Streamlit download."""
    zip_buffer = BytesIO()
    frames = build_output_frames(results_df)

    with ZipFile(zip_buffer, mode="w", compression=ZIP_DEFLATED) as archive:
        for filename, frame in frames.items():
            archive.writestr(filename, dataframe_to_excel_bytes(frame).getvalue())

    zip_buffer.seek(0)
    return zip_buffer


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Transactions") -> BytesIO:
    """Write a DataFrame to a professionally formatted Excel workbook."""
    buffer = BytesIO()
    export_df = df.copy()
    export_df = export_df.fillna("")

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        _format_worksheet(worksheet, export_df)

    buffer.seek(0)
    return buffer


def _ensure_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.reindex(columns=OUTPUT_COLUMNS)


def _qbo_signed_amount(row: pd.Series) -> float:
    spent = _safe_number(row.get("QBO Spent", 0))
    received = _safe_number(row.get("QBO Received", 0))

    if spent:
        return -abs(spent)
    if received:
        return abs(received)
    return 0


def _safe_number(value) -> float:
    if value is None:
        return 0
    if isinstance(value, str) and not value.strip():
        return 0

    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return 0
    return float(number)


def _safe_excel_text(value) -> str:
    if value is None:
        return ""
    if not pd.api.types.is_scalar(value):
        return str(value)

    try:
        is_missing = pd.isna(value)
    except (TypeError, ValueError):
        return str(value)

    try:
        if bool(is_missing):
            return ""
    except (TypeError, ValueError):
        return str(value)

    return str(value)


def _format_worksheet(worksheet, df: pd.DataFrame) -> None:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    if worksheet.dimensions and worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        values = [_safe_excel_text(column_cells[0].value)]

        # Convert safely before measuring because Excel cells can hold blank or unusual values.
        for cell in column_cells[1:]:
            values.append(_safe_excel_text(cell.value))

        max_len = max((len(value) for value in values), default=0)
        worksheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 48)

        lower_header = values[0].lower()
        for cell in column_cells[1:]:
            if "date" in lower_header:
                cell.number_format = "yyyy-mm-dd"
            elif _is_amount_column(lower_header):
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'


def _is_amount_column(header: str) -> bool:
    return any(token in header for token in ["amount", "spent", "received"])
